from __future__ import annotations

import json
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

from bybit_p2p._exceptions import FailedRequestError

from services.order_processing.messaging import counterparty_realname, country_name
from services.order_processing.payments import extract_iban, extract_pl_phone, extract_pln_payment_buy
from services.orders_service import _fetch_counterparty_info

CHAT_PAGE_SIZE = 200
ORDERS_PAGE_SIZE = 50


def _parse_ms(value: Any) -> Optional[int]:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _extract_order_items(response: Dict[str, Any]) -> List[Dict[str, Any]]:
    if not isinstance(response, dict):
        return []
    result = response.get("result")
    if not isinstance(result, dict):
        return []
    items = result.get("items")
    return items if isinstance(items, list) else []


def _extract_order_count(response: Dict[str, Any]) -> Optional[int]:
    if not isinstance(response, dict):
        return None
    result = response.get("result")
    if not isinstance(result, dict):
        return None
    count = result.get("count")
    try:
        return int(count)
    except (TypeError, ValueError):
        return None


def _get_orders_with_fallback(
    api,
    params: Dict[str, Any],
) -> tuple[Dict[str, Any], Optional[str]]:
    try:
        return api.get_orders(**params), None
    except FailedRequestError as exc:
        if exc.status_code != 40001:
            raise

    if "status" in params:
        fallback = dict(params)
        fallback.pop("status", None)
        try:
            return api.get_orders(**fallback), str(params["status"])
        except FailedRequestError as exc:
            if exc.status_code != 40001:
                raise

    for key in ("tokenId", "side"):
        if key not in params:
            continue
        fallback = dict(params)
        fallback.pop(key, None)
        try:
            return api.get_orders(**fallback), None
        except FailedRequestError as exc:
            if exc.status_code != 40001:
                raise

    return api.get_orders(**params), None


def _extract_chat_items(response: Dict[str, Any]) -> List[Dict[str, Any]]:
    if not isinstance(response, dict):
        return []
    result = response.get("result")
    if isinstance(result, dict):
        items = result.get("result") or result.get("items") or []
    else:
        items = []
    return items if isinstance(items, list) else []


def _normalize_side(value: Any) -> str:
    text = str(value).lower()
    if text in {"1", "sell"}:
        return "SELL"
    return "BUY"


def _normalize_status(value: Any) -> str:
    status = _parse_ms(value)
    if status == 40:
        return "Cancelled"
    if status == 50:
        return "Completed"
    return str(value or "")


def _clean_missing(value: str) -> str:
    if not value or str(value).strip().lower() == "not found":
        return ""
    return str(value)


def _format_kyc_country(counterparty: Dict[str, Any]) -> str:
    code = (counterparty or {}).get("kycCountryCode") or ""
    if not code:
        return ""
    name = country_name(code)
    if name:
        return f"{code}, {name}"
    return str(code)


def _parse_date_input(value: Any, *, is_end: bool) -> datetime:
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        boundary = time.max if is_end else time.min
        dt = datetime.combine(value, boundary)
    elif isinstance(value, (int, float)):
        seconds = value / 1000 if value > 10**12 else value
        dt = datetime.fromtimestamp(seconds)
    elif isinstance(value, str):
        raw = value.strip()
        dt = datetime.fromisoformat(raw)
        if "T" not in raw and " " not in raw:
            boundary = time.max if is_end else time.min
            dt = dt.replace(
                hour=boundary.hour,
                minute=boundary.minute,
                second=boundary.second,
                microsecond=boundary.microsecond,
            )
    else:
        raise TypeError("Unsupported date input")
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def _format_created_at(value: Any) -> str:
    if value is None or value == "":
        return ""
    dt: Optional[datetime] = None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, (int, float)):
        seconds = value / 1000 if value > 10**12 else value
        dt = datetime.fromtimestamp(seconds)
    elif isinstance(value, str):
        raw = value.strip()
        if raw.isdigit():
            digits = int(raw)
            seconds = digits / 1000 if digits > 10**12 else digits
            dt = datetime.fromtimestamp(seconds)
        else:
            try:
                dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            except ValueError:
                dt = None
    if not dt:
        return ""
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _to_ms(dt: datetime) -> int:
    return int(dt.timestamp() * 1000)


def _to_seconds(dt: datetime) -> int:
    return int(dt.timestamp())


def _payment_method_label(order: Dict[str, Any]) -> str:
    terms = order.get("paymentTermList") or order.get("paymentTerms") or []
    name = ""
    payment_type = ""
    if isinstance(terms, list) and terms:
        term = terms[0]
        config = term.get("paymentConfigVo") or {}
        name = str(config.get("paymentName") or "") or ""
        payment_type = str(term.get("paymentType") or config.get("paymentType") or "")
    else:
        payment_type = str(order.get("paymentType") or "")
    if not name and not payment_type:
        return ""
    return json.dumps([name, payment_type], ensure_ascii=False)


def _get_my_account_id(api) -> str:
    try:
        resp = api.get_account_information()
    except Exception:
        return ""
    result = resp.get("result") if isinstance(resp, dict) else None
    return str(result.get("accountId") or "") if isinstance(result, dict) else ""


def _collect_chat_contacts(api, order_id: str, my_account_id: str) -> Tuple[set[str], set[str], str]:
    ibans: set[str] = set()
    phones: set[str] = set()
    last_phone = ""
    last_phone_id: Optional[int] = None
    start_message_id: Optional[int] = None
    while True:
        params: Dict[str, Any] = {"orderId": order_id, "size": str(CHAT_PAGE_SIZE)}
        if start_message_id is not None:
            params["startMessageId"] = str(start_message_id)
        try:
            resp = api.get_chat_messages(**params)
        except Exception:
            break
        items = _extract_chat_items(resp)
        if not items:
            break
        max_id = start_message_id
        for item in items:
            account_id = str(item.get("accountId") or "")
            if account_id and account_id == my_account_id:
                continue
            if str(item.get("contentType") or "str") != "str":
                continue
            msg = str(item.get("message") or "")
            if not msg:
                continue
            msg_id = _parse_ms(item.get("id"))
            iban = extract_iban(msg)
            if iban:
                ibans.add(iban)
            phone = extract_pl_phone(msg)
            if phone:
                phones.add(phone)
                if msg_id is not None and (last_phone_id is None or msg_id > last_phone_id):
                    last_phone_id = msg_id
                    last_phone = phone
            if msg_id is not None:
                max_id = max(max_id or msg_id, msg_id)
        if max_id is None or max_id == start_message_id:
            break
        if len(items) < CHAT_PAGE_SIZE:
            break
        start_message_id = max_id
    return ibans, phones, last_phone


def _merge_contacts(existing: str, additions: Iterable[str]) -> str:
    existing_set = {part.strip() for part in str(existing).split(",") if part.strip()}
    merged = sorted(set(existing_set) | set(additions))
    return ", ".join(merged) if merged else ""


def export_bybit_history_excels(
    api,
    start_date: Any,
    end_date: Any,
    *,
    time_unit: str = "ms",
    chunk_days: int = 90,
    page_size: int = 10,
    status: int | None = None,
    token_id: str | None = None,
    currency_id: str | None = None,
    side: int | None = None,
    output_path: str | None = None,
) -> Path:
    start_dt = _parse_date_input(start_date, is_end=False)
    end_dt = _parse_date_input(end_date, is_end=True)
    orders: List[Dict[str, Any]] = []
    total_count: Optional[int] = None
    current_from = start_dt
    while current_from < end_dt:
        current_to = min(current_from + timedelta(days=chunk_days), end_dt)
        if time_unit == "ms":
            start_ts = _to_ms(current_from)
            end_ts = _to_ms(current_to)
        else:
            start_ts = _to_seconds(current_from)
            end_ts = _to_seconds(current_to)

        page = 1
        while True:
            params: Dict[str, Any] = {
                "page": page,
                "size": page_size,
                "beginTime": start_ts,
                "endTime": end_ts,
            }
            if status is not None:
                params["status"] = str(status)
            if token_id:
                params["tokenId"] = str(token_id)
            if currency_id:
                params["currencyId"] = str(currency_id)
            if side is not None:
                params["side"] = str(side)
            resp, filter_status = _get_orders_with_fallback(api, params)
            if total_count is None:
                total_count = _extract_order_count(resp)
            batch = _extract_order_items(resp)
            if filter_status is not None:
                batch = [
                    item
                    for item in batch
                    if str(item.get("status") or "") == filter_status
                ]
            if not batch:
                break
            orders.extend(batch)
            print(f"[orders] total={len(orders)}")
            if len(batch) < page_size:
                break
            page += 1
        current_from = current_to

    my_account_id = _get_my_account_id(api)
    rows: List[List[Any]] = []
    with tqdm(total=total_count, unit="order") as progress:
        for order in orders:
            order_id = str(order.get("id") or order.get("orderId") or "")
            details: Dict[str, Any] = {}
            if order_id:
                try:
                    resp = api.get_order_details(orderId=order_id)
                    details = resp.get("result") if isinstance(resp, dict) else {}
                except Exception:
                    details = {}
            record = details if isinstance(details, dict) and details else order
            counterparty = record.get("counterparty_info") or _fetch_counterparty_info(api, record) or {}

            side = _normalize_side(record.get("side"))
            payment_method = _payment_method_label(record)
            account_no = ""
            phone = ""
            last_chat_phone = ""
            if side == "BUY":
                buy_info = extract_pln_payment_buy(record)
                account_no = _clean_missing(buy_info.get("iban", ""))
                phone = _clean_missing(buy_info.get("phone", ""))
                if order_id:
                    chat_ibans, chat_phones, last_chat_phone = _collect_chat_contacts(
                        api, order_id, my_account_id
                    )
                    account_no = _merge_contacts(account_no, chat_ibans)
                    phone = _merge_contacts(phone, chat_phones)
            elif order_id:
                _, _, last_chat_phone = _collect_chat_contacts(api, order_id, my_account_id)

            if last_chat_phone:
                phone = last_chat_phone

            rows.append(
                [
                    order_id,
                    side,
                    record.get("amount"),
                    record.get("currencyId"),
                    record.get("price"),
                    record.get("quantity"),
                    record.get("tokenId"),
                    record.get("targetNickName"),
                    _normalize_status(record.get("status")),
                    _format_created_at(record.get("createDate")),
                    counterparty_realname(record, counterparty),
                    _format_kyc_country(counterparty),
                    payment_method,
                    account_no,
                    phone,
                ]
            )
            progress.update(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bybit Orders"
    buy_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    sell_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ws.append(
        [
            "Order No.",
            "Side",
            "Fiat Amount",
            "Currency",
            "Price",
            "Coin Amount",
            "Cryptocurrency",
            "Counterparty",
            "Status",
            "Created At",
            "Counterparty Real Name",
            "KYC Country",
            "Payment Method",
            "Account No.",
            "Phone",
        ]
    )
    for row in rows:
        ws.append(row)
        side = str(row[1]).upper()
        if side in {"BUY", "SELL"}:
            fill = buy_fill if side == "BUY" else sell_fill
            for cell in ws[ws.max_row]:
                cell.fill = fill

    target = Path(output_path) if output_path else Path("playground_results")
    if target.suffix.lower() != ".xlsx":
        target.mkdir(parents=True, exist_ok=True)
        name = f"bybit_history_{start_dt:%Y%m%d}_{end_dt:%Y%m%d}.xlsx"
        path = target / name
    else:
        target.parent.mkdir(parents=True, exist_ok=True)
        path = target
    wb.save(path)
    return path
